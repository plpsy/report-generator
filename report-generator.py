import tkinter
from tkinter import ttk,filedialog,messagebox,scrolledtext
from tkinter import *
import openpyxl
import threading
import time
import serial
import serial.tools.list_ports
import re
import os
from code_dic import MY_CODE
from docx import Document
from tkinter import ttk
import sv_ttk
import gen

class MY_GUI():

    #构造函数
    def __init__(self,name):
        self.init_window_name=name
    #窗口控件设置初始化
    def set_init_window(self):
        self.init_window_name.title('测试报告自动生成工具')
        self.init_window_name.geometry('940x500')
        self.init_window_name.iconphoto(True, tkinter.PhotoImage(file=os.path.join(os.getcwd(), 'logo.png')))
        #串口选择框架内部标签
        self.com_label=ttk.Label(text='COMx: ')
        self.com_label.place(x=20, y=20)   
        self.baudrate_label=ttk.Label(text='Baudrate: ')
        self.baudrate_label.place(x=20, y=60) 
        #串口框架内部下拉选项框
        self.com_choose=StringVar()
        self.com_choose_combo=ttk.Combobox(width=20, textvariable=self.com_choose)
        self.com_choose_combo['state']='readonly'
        self.com_choose_combo.place(x=100,y=10)
        self.com_choose_combo['values']=self.com_name_get()
        self.com_choose_combo.current(0)
        #波特率选项框
        self.baudrate_value=StringVar(value='115200')
        self.baudrate_choose_combo=ttk.Combobox(width=20,textvariable=self.baudrate_value)
        self.baudrate_choose_combo['values']=('9600', '115200')
        self.baudrate_choose_combo['state']='readonly'
        self.baudrate_choose_combo.place(x=100,y=50)		
        #串口框架内部按钮
        self.connect_button=ttk.Button(text='连接' ,width=10, command=self.com_connect)
        self.connect_button.place(x=280,y=10)	
        self.cancel_button=ttk.Button(text='取消' ,width=10, command=self.com_cancel)
        self.cancel_button.place(x=280,y=50)	
        self.com_log_text=Text(width=50, height=5)
        self.com_log_text.place(x=20, y=90)
        self.com_log_text.insert(END, '串口连接状态'+'\n')

        self.file_choose_button=ttk.Button(self.init_window_name, text='选择指令文件', width=13,command=self.thread_file)
        self.file_choose_button.place(x=20, y=170)
        self.file_path_text=Text(self.init_window_name, width=35, height=1)
        self.file_path_text.place(x=120,y=175)
		
        self.num_input_label=ttk.Label(self.init_window_name, text='输入指令：')
        self.num_input_label.place(x=20, y=460) 
        self.input_num=StringVar()
        self.input_num_entry=Entry(self.init_window_name, textvariable=self.input_num, width=25)
        self.input_num_entry.place(x=90, y=460) 
        self.conduct_button=ttk.Button(self.init_window_name, text='执行', width=4, command=self.com_output)
        self.conduct_button.place(x=250, y=455)        
        self.clear_button=ttk.Button(self.init_window_name, text='清空', width=5, command=self.thread_clear)
        self.clear_button.place(x=310, y=455)        
        
        # self.result_data_label=ttk.Label(self.init_window_name, text='串口输出')
        # self.result_data_label.place(x=400, y=20)

        self.file_choose_button1=ttk.Button(self.init_window_name, text='选择串口文件', width=13,command=self.thread_file1)
        self.file_choose_button1.place(x=400, y=15)
        self.file_path_text1=Text(self.init_window_name, width=40, height=1)
        self.file_path_text1.place(x=500,y=15)
        self.log_save_button=ttk.Button(self.init_window_name,text='生成测试报告', width=12, command=self.thread_save)
        self.log_save_button.place(x=800, y=15)
        #处理结果显示滚动文本框
        self.result_text=scrolledtext.ScrolledText(self.init_window_name, width=70, height=32)
        self.result_text.place(x=400, y=50)

        #代码解析后进行显示
        self.code_frame=Frame(self.init_window_name, width=78, height=29, bg='white')
        self.code_frame.place(x=20,y=210)
        #解析后的代码放在表格内显示
        self.code_tree=ttk.Treeview(self.code_frame,show='headings', height=10, columns=('0','1','2'))
        self.code_bar=ttk.Scrollbar(self.code_frame,orient=VERTICAL,command=self.code_tree.yview)
        self.code_tree.configure(yscrollcommand=self.code_bar.set)
        self.code_tree.grid(row=0,column=0,sticky=NSEW)
        self.code_bar.grid(row=0,column=1,sticky=NS)
        self.code_tree.column('0',width=30)
        self.code_tree.column('1',width=250)
        self.code_tree.column('2',width=50)
        self.code_tree.heading('0',text='序号')
        self.code_tree.heading('1',text='命令')
        self.code_tree.heading('2',text='备注')
        self.code_tree.bind("<<TreeviewSelect>>", self.codetree_select)

        self.output_lines = []
        self.read_thread_runnning = False

    def codetree_select(self, event):
        curItem = self.code_tree.focus()
        out_str = self.code_tree.item(curItem)["values"][1]        
        self.input_num_entry.delete('0', END)
        self.input_num_entry.insert('0', out_str)

    def start_read_thread(self):
        if(self.read_thread_runnning):
            return
        thisthread = threading.Thread(target=self.read_thread)
        thisthread.start()
    
    def read_thread(self):
        try:
            self.read_thread_runnning = True
            self.ser=serial.Serial(self.ser_name)
            self.ser.baudrate=self.ser_baudrate
            self.ser.timeout=0.5
            self.com_log_text.insert(END,time.ctime(time.time())+'\t\t'+'串口成功打开'+'\n')
            self.com_log_text.see(tkinter.END)
            self.com_log_text.update()
            while self.read_thread_runnning:
                data = self.ser.readline()#字节类型
                if not data:
                    continue
                lines = data.decode('utf-8').split("\r")
                for line in lines:
                    self.output_lines.append(line)
                    self.result_text.insert(END, line + '\r\n')
                    self.result_text.see(tkinter.END)
                self.result_text.update()
        except:
            newline=time.ctime(time.time())+'\t\t'+'串口打开故障或串口被关闭'+'\n'
            self.com_log_text.insert('END',newline)
            self.com_log_text.see(tkinter.END)
            self.com_log_text.update()

            self.read_thread_runnning = False
    #自动获取当前连接的串口名
    # def com_name_get(self):
    #     self.port_list=list(serial.tools.list_ports.comports())
    #     self.com_port_names=[]
    #     self.pattern=re.compile(r'[(](.*?)[)]',re.S)
    #     if len(self.port_list)>0:
    #         for i in range(len(self.port_list)):
    #             self.com_name=re.findall(self.pattern,str(self.port_list[i]))
    #             self.com_port_names.append(self.com_name)
    #     return self.com_port_names
    def com_name_get(self):
        self.port_list=list(serial.tools.list_ports.comports())
        self.com_port_names=[p[0] for p in self.port_list]

        return self.com_port_names

    #连接按键的执行内容
    def com_connect(self):
        # self.result_text.insert(END,'请连接串口设备'+'\n')
        self.ser_name=str(self.com_choose.get())
        self.ser_baudrate=str(self.baudrate_value.get())
        self.start_read_thread()

    #取消按键的执行内容
    def com_cancel(self):
        self.result_text.delete('1.0','end')
        self.read_thread_runnning = False
        try:
            self.ser.close()
            newline=time.ctime(time.time())+'\t\t'+'串口被关闭'+'\n'
        except:
            newline=time.ctime(time.time())+'\t\t'+'串口未打开'+'\n'
        self.output_lines = []
        self.com_log_text.insert(END,newline)
        self.com_log_text.see(tkinter.END)
        self.com_log_text.update()

    #执行按键的执行内容
    def com_output(self):
        if self.input_num_entry.get():
            self.testnum = self.input_num_entry.get()
            out_str = str(self.testnum) + '\r\n'
            try:
                self.ser.write(out_str.encode())
            except:
                tkinter.messagebox.showerror('执行异常','发送指令失败')

    #执行按键的执行内容
    def com_output1(self):
        autodocument=open('autoresult.txt','a')
        self.code_tree_items=self.code_tree.get_children()
        for i in range(self.codeline_counter):
            self.code_tree.set(self.code_tree_items[i],column=2,value=' ')
        self.code_result_sheet=['failed'for i in range(self.codeline_counter)]
        self.functionname=[]
        if self.input_num_entry.get():
            self.testnum=self.input_num_entry.get()
        else:
            self.testnum='未标'
        newline='编号: '+'\t'+str(self.testnum)
        print(newline,file=autodocument)
        try:
            for i in range(self.codeline_counter):
                function=str(self.code_sheet[i][1])
                self.functionandparam=re.split(',',function)                              #用，将函数名与参数分开
                functionname=self.functionandparam[0]                                     #第一个元素取为函数名，后面的都是参数
                self.functionname.append(functionname)
                my_code=MY_CODE(self.ser,self.result_text,self.functionandparam)
                result=my_code.function_choose_do(functionname)
                if result ==1:
                    self.code_tree.set(self.code_tree_items[i],column=2,value='ok')
                    self.code_result_sheet[i]='ok'
                else:
                    self.code_tree.set(self.code_tree_items[i],column=2,value='failed')
                    self.code_result_sheet[i]='failed'
                newline=str(self.code_sheet[i][1])+':'+'\t'+str(self.code_result_sheet[i])
                print(newline,file=autodocument)
        except:
            tkinter.messagebox.showerror('执行异常','代码执行异常，请仔细检查代码及格式！')
        failnum = 0
        for code_result in self.code_result_sheet:
            if code_result == 'ok':
                failnum += 1
            else:
                break
        if failnum < self.codeline_counter :
            self.isok = '失败'
            self.failcode=str(self.code_sheet[failnum][1])
        else:
            self.isok='成功'
            self.failcode = ' '
        self.result_tree.insert('','end',values=[self.testnum,self.isok,self.failcode])

    #新建线程，负责选择代码文件、保存代码执行结果和清空代码表格

    #新建选择文件线程
    def thread_file(self):
        thisthread=threading.Thread(target=self.file_choose)
        thisthread.start()

    #选择文件打开，并在界面中显示
    def file_choose(self):
        self.codeline_counter=0
        self.root=Tk()
        self.root.withdraw()
        file_path=filedialog.askopenfilename()
        self.file_path_text.insert(END,file_path)
        wb=openpyxl.load_workbook(file_path)
        sheet=wb['Sheet1']
        self.code_sheet=[[0 for i in range(3)]for j in range(20)]
        for i in range(20):
            if sheet.cell(row=i+2,column=1).value:
                self.codeline_counter +=1
                self.code_context=[]
                for j in range(3):
                    self.code_sheet[i][j]=sheet.cell(row=i+2,column=j+1).value
                    self.code_context.append(self.code_sheet[i][j])
                self.code_tree.insert('', i, values=self.code_context)  #code_tree用来将代码规则的显示，code_sheet是一个数组，便于取值运算

    #新建选择文件线程
    def thread_file1(self):
        thisthread=threading.Thread(target=self.file_choose1)
        thisthread.start()

    #选择文件打开，并在界面中显示
    def file_choose1(self):
        self.root=Tk()
        self.root.withdraw()
        file_path=filedialog.askopenfilename()
        self.file_path_text1.insert(END,file_path)
        self.output_lines = gen.get_output(file_path)

    #新建线程保存执行结果
    def thread_save(self):
        thisthread=threading.Thread(target=self.save_to_docx)
        thisthread.start()

    def save_to_docx(self):

        file_path = filedialog.asksaveasfilename(initialfile='测试记录.docx',filetypes=[("word文件", ".docx")])
        self.output_lines = [x.lower() for x in self.output_lines]
        # for line in self.output_lines:
        #     print(line)
        gen.generate_docx(self.output_lines, file_path)

    #保存代码执行结果日志
    def code_log_save(self):
        document=open('result.txt','a')#a表示写入时不将原有内容清除
        try:
            firstline='编号： '+str(self.testnum)
        except:
            firstline='编号： '+'未标注'
        print(firstline,file=document)
        secondline='指令'.rjust(10,' ')+'\t'+'执行结果'
        print(secondline,file=document)
        try:
            for i in range(self.codeline_counter):
                print(str(self.functionname[i]).rjust(12,' ')+': '+'\t'+str(self.code_result_sheet[i]),file=document)
            tkinter.messagebox.showinfo('保存成功','Saved successfully!')
        except:
            thirdline='\t\t'+'代码未执行'
            print(thirdline,file=document)
            tkinter.messagebox.showinfo('异常','没有选择文件或代码执行异常')

    #新建线程清空所选文件以备重新选择
    def thread_clear(self):
        thisthread=threading.Thread(target=self.file_clear)
        thisthread.start()

    #删除所选文件，清空解析后的代码表格
    def file_clear(self):
        self.file_path_text.delete('1.0','end')
        self.input_num.set(' ')
        code_items=self.code_tree.get_children()
        for item in code_items:
            self.code_tree.delete(item)

#主线程
def start():
    init_window=tkinter.Tk()
    my_window=MY_GUI(init_window)
    my_window.set_init_window()

#    sv_ttk.use_dark_theme()
#    sv_ttk.use_light_theme()	
#    sv_ttk.set_theme("light")
    init_window.mainloop()

start()

